# app/handlers/admin_extended.py
"""Расширенные админские функции: рассылка, поиск, экспорт, массовые операции"""

import io
import json
import logging
from datetime import datetime
from typing import List, Set

from aiogram import types
from aiogram.dispatcher import FSMContext
from functools import wraps

from app.constants import (
    BTN_BROADCAST, BTN_SEARCH, BTN_MASS_OPS, BTN_EXPORT,
    CB_BROADCAST, CB_BROADCAST_ALL, CB_BROADCAST_SELECT, CB_BC_MANAGER,
    CB_BC_CONFIRM, CB_BC_CANCEL, CB_BC_ADD_PHOTO, CB_BC_SKIP_PHOTO, CB_BC_DONE,
    CB_SEARCH,
    CB_MASS_OPS, CB_MASS_ARCHIVE, CB_MASS_DELETE, CB_MASS_CONFIRM, CB_MASS_CANCEL,
    CB_EXPORT_EXCEL, CB_EXPORT_PDF,
)
from app.db import (
    get_mode, get_user_by_tgid,
    get_all_active_managers, get_managers_by_ids,
    search_requests, list_all_requests,
    mass_archive_requests, mass_delete_requests,
    get_stats_for_export, get_managers_stats_for_export,
    get_requests_by_status, get_overall_stats, log_activity,
)
from app.keyboards import (
    broadcast_start_inline, broadcast_managers_select_inline,
    broadcast_confirm_inline, broadcast_photo_inline,
    search_results_inline,
    mass_ops_start_inline, mass_ops_confirm_inline,
    export_options_inline, admin_main_inline,
)
from app.states import AdminBroadcast, SearchRequest
from app.utils import e

log = logging.getLogger("bot")


def require_admin(handler):
    """Декоратор для проверки прав администратора"""
    @wraps(handler)
    async def wrapper(obj, *args, **kwargs):
        if isinstance(obj, types.Message):
            user_id = obj.from_user.id
        else:
            user_id = obj.from_user.id

        if get_mode(user_id) != "admin":
            if isinstance(obj, types.CallbackQuery):
                return await obj.answer("⛔ Требуются права администратора", show_alert=True)
            return await obj.answer("⛔ Доступ запрещён.")

        return await handler(obj, *args, **kwargs)
    return wrapper


def register(dp, bot):
    """Регистрация расширенных админских обработчиков"""

    # ==================== РАССЫЛКА ====================

    @require_admin
    async def cmd_broadcast(message: types.Message, state: FSMContext):
        """Начало рассылки"""
        await state.finish()  # Сбрасываем предыдущее состояние

        managers = get_all_active_managers()
        if not managers:
            return await message.answer("Нет активных менеджеров для рассылки.")

        await message.answer(
            f"📢 <b>Рассылка сообщений</b>\n\n"
            f"Активных менеджеров: <b>{len(managers)}</b>\n\n"
            "Выберите получателей:",
            reply_markup=broadcast_start_inline()
        )

    dp.register_message_handler(cmd_broadcast, lambda m: m.text == BTN_BROADCAST, state="*")
    dp.register_message_handler(cmd_broadcast, commands=["broadcast"], state="*")

    @require_admin
    async def cb_broadcast_menu(call: types.CallbackQuery, state: FSMContext):
        """Возврат к меню рассылки"""
        await call.answer()
        await state.finish()

        managers = get_all_active_managers()
        await call.message.edit_text(
            f"📢 <b>Рассылка сообщений</b>\n\n"
            f"Активных менеджеров: <b>{len(managers)}</b>\n\n"
            "Выберите получателей:",
            reply_markup=broadcast_start_inline()
        )

    dp.register_callback_query_handler(cb_broadcast_menu, lambda c: c.data == CB_BROADCAST, state="*")

    @require_admin
    async def cb_broadcast_all(call: types.CallbackQuery, state: FSMContext):
        """Рассылка всем менеджерам"""
        await call.answer()

        managers = get_all_active_managers()
        manager_ids = [str(m['id']) for m in managers]
        tg_ids = [m['tg_id'] for m in managers]

        await state.update_data(
            broadcast_manager_ids=manager_ids,
            broadcast_tg_ids=tg_ids,
            broadcast_count=len(managers)
        )

        await AdminBroadcast.composing_message.set()
        await call.message.edit_text(
            f"📢 <b>Рассылка всем ({len(managers)} чел.)</b>\n\n"
            "Отправьте сообщение, которое будет разослано.\n"
            "Можете отправить текст или фото с подписью.\n\n"
            "Для отмены напишите /cancel"
        )

    dp.register_callback_query_handler(cb_broadcast_all, lambda c: c.data == CB_BROADCAST_ALL, state="*")

    @require_admin
    async def cb_broadcast_select(call: types.CallbackQuery, state: FSMContext):
        """Выбор получателей"""
        await call.answer()

        managers = get_all_active_managers()
        await state.update_data(broadcast_selected=set())
        await AdminBroadcast.selecting_recipients.set()

        await call.message.edit_text(
            f"👤 <b>Выберите получателей</b>\n\n"
            f"Всего менеджеров: {len(managers)}\n"
            "Нажмите на имя, чтобы выбрать/убрать:",
            reply_markup=broadcast_managers_select_inline(managers, set())
        )

    dp.register_callback_query_handler(cb_broadcast_select, lambda c: c.data == CB_BROADCAST_SELECT, state="*")

    @require_admin
    async def cb_toggle_manager(call: types.CallbackQuery, state: FSMContext):
        """Переключение выбора менеджера"""
        manager_id = call.data[len(CB_BC_MANAGER):]

        data = await state.get_data()
        selected: Set[str] = set(data.get('broadcast_selected', []))

        if manager_id in selected:
            selected.discard(manager_id)
        else:
            selected.add(manager_id)

        await state.update_data(broadcast_selected=list(selected))

        managers = get_all_active_managers()
        await call.message.edit_reply_markup(
            reply_markup=broadcast_managers_select_inline(managers, selected)
        )
        await call.answer(f"Выбрано: {len(selected)}")

    dp.register_callback_query_handler(
        cb_toggle_manager,
        lambda c: c.data and c.data.startswith(CB_BC_MANAGER),
        state=AdminBroadcast.selecting_recipients
    )

    @require_admin
    async def cb_broadcast_done_select(call: types.CallbackQuery, state: FSMContext):
        """Завершение выбора получателей"""
        data = await state.get_data()
        selected = data.get('broadcast_selected', [])

        if not selected:
            return await call.answer("Выберите хотя бы одного получателя", show_alert=True)

        # Получаем tg_id выбранных
        managers = get_managers_by_ids(list(selected))
        tg_ids = [m['tg_id'] for m in managers]

        await state.update_data(
            broadcast_manager_ids=list(selected),
            broadcast_tg_ids=tg_ids,
            broadcast_count=len(selected)
        )

        await AdminBroadcast.composing_message.set()
        await call.message.edit_text(
            f"📢 <b>Рассылка ({len(selected)} чел.)</b>\n\n"
            "Отправьте сообщение для рассылки.\n"
            "Можете отправить текст или фото с подписью.\n\n"
            "Для отмены напишите /cancel"
        )

    dp.register_callback_query_handler(
        cb_broadcast_done_select,
        lambda c: c.data == CB_BC_DONE,
        state=AdminBroadcast.selecting_recipients
    )

    @require_admin
    async def process_broadcast_message(message: types.Message, state: FSMContext):
        """Получение сообщения для рассылки"""
        if message.text and message.text.startswith('/'):
            if message.text == '/cancel':
                await state.finish()
                return await message.answer("Рассылка отменена.", reply_markup=admin_main_inline())

        data = await state.get_data()
        count = data.get('broadcast_count', 0)

        # Сохраняем сообщение
        if message.photo:
            await state.update_data(
                broadcast_text=message.caption or "",
                broadcast_photo=message.photo[-1].file_id,
                broadcast_has_photo=True
            )
        else:
            await state.update_data(
                broadcast_text=message.text or "",
                broadcast_photo=None,
                broadcast_has_photo=False
            )

        # Переход к подтверждению
        await AdminBroadcast.confirming.set()

        preview_text = (message.caption or message.text or "")[:200]
        if len(message.caption or message.text or "") > 200:
            preview_text += "..."

        has_photo = "📷 С фото" if message.photo else "📝 Только текст"

        await message.answer(
            f"📢 <b>Подтверждение рассылки</b>\n\n"
            f"Получателей: <b>{count}</b>\n"
            f"Тип: {has_photo}\n\n"
            f"<b>Превью:</b>\n{e(preview_text)}",
            reply_markup=broadcast_confirm_inline(bool(message.photo))
        )

    dp.register_message_handler(
        process_broadcast_message,
        content_types=['text', 'photo'],
        state=AdminBroadcast.composing_message
    )

    @require_admin
    async def cb_add_photo(call: types.CallbackQuery, state: FSMContext):
        """Добавление фото к рассылке"""
        await call.answer()
        await AdminBroadcast.adding_photo.set()

        await call.message.edit_text(
            "📷 <b>Добавление фото</b>\n\n"
            "Отправьте фото для рассылки:",
            reply_markup=broadcast_photo_inline()
        )

    dp.register_callback_query_handler(cb_add_photo, lambda c: c.data == CB_BC_ADD_PHOTO, state=AdminBroadcast.confirming)

    @require_admin
    async def process_broadcast_photo(message: types.Message, state: FSMContext):
        """Получение фото для рассылки"""
        if not message.photo:
            return await message.answer("Отправьте фото.")

        await state.update_data(
            broadcast_photo=message.photo[-1].file_id,
            broadcast_has_photo=True
        )

        data = await state.get_data()
        count = data.get('broadcast_count', 0)
        preview_text = data.get('broadcast_text', '')[:200]

        await AdminBroadcast.confirming.set()
        await message.answer(
            f"📢 <b>Подтверждение рассылки</b>\n\n"
            f"Получателей: <b>{count}</b>\n"
            f"Тип: 📷 С фото\n\n"
            f"<b>Текст:</b>\n{e(preview_text)}",
            reply_markup=broadcast_confirm_inline(True)
        )

    dp.register_message_handler(
        process_broadcast_photo,
        content_types=['photo'],
        state=AdminBroadcast.adding_photo
    )

    @require_admin
    async def cb_skip_photo(call: types.CallbackQuery, state: FSMContext):
        """Пропуск фото"""
        await call.answer()

        data = await state.get_data()
        count = data.get('broadcast_count', 0)
        preview_text = data.get('broadcast_text', '')[:200]

        await AdminBroadcast.confirming.set()
        await call.message.edit_text(
            f"📢 <b>Подтверждение рассылки</b>\n\n"
            f"Получателей: <b>{count}</b>\n"
            f"Тип: 📝 Только текст\n\n"
            f"<b>Текст:</b>\n{e(preview_text)}",
            reply_markup=broadcast_confirm_inline(False)
        )

    dp.register_callback_query_handler(cb_skip_photo, lambda c: c.data == CB_BC_SKIP_PHOTO, state=AdminBroadcast.adding_photo)

    @require_admin
    async def cb_broadcast_confirm(call: types.CallbackQuery, state: FSMContext):
        """Отправка рассылки"""
        await call.answer("Отправка...")

        data = await state.get_data()
        tg_ids = data.get('broadcast_tg_ids', [])
        text = data.get('broadcast_text', '')
        photo = data.get('broadcast_photo')

        success = 0
        failed = 0

        for tg_id in tg_ids:
            try:
                if photo:
                    await bot.send_photo(tg_id, photo, caption=text)
                else:
                    await bot.send_message(tg_id, text)
                success += 1
            except Exception as e:
                log.warning("Broadcast failed to %s: %s", tg_id, e)
                failed += 1

        await state.finish()

        # Логируем
        admin = get_user_by_tgid(call.from_user.id)
        if admin:
            log_activity(str(admin['id']), "broadcast_sent", None, None, {
                "recipients": len(tg_ids),
                "success": success,
                "failed": failed
            })

        await call.message.edit_text(
            f"✅ <b>Рассылка завершена</b>\n\n"
            f"📤 Отправлено: {success}\n"
            f"❌ Ошибок: {failed}",
            reply_markup=admin_main_inline()
        )

    dp.register_callback_query_handler(cb_broadcast_confirm, lambda c: c.data == CB_BC_CONFIRM, state=AdminBroadcast.confirming)

    @require_admin
    async def cb_broadcast_cancel(call: types.CallbackQuery, state: FSMContext):
        """Отмена рассылки"""
        await call.answer("Отменено")
        await state.finish()
        await call.message.edit_text("Рассылка отменена.", reply_markup=admin_main_inline())

    dp.register_callback_query_handler(cb_broadcast_cancel, lambda c: c.data == CB_BC_CANCEL, state="*")

    # ==================== ПОИСК ====================

    @require_admin
    async def cmd_search(message: types.Message, state: FSMContext):
        """Начало поиска"""
        await SearchRequest.waiting_query.set()
        await message.answer(
            "🔍 <b>Поиск заявок</b>\n\n"
            "Введите название компании, имя клиента или сферу деятельности:\n\n"
            "Для отмены напишите /cancel"
        )

    dp.register_message_handler(cmd_search, lambda m: m.text == BTN_SEARCH, state="*")
    dp.register_message_handler(cmd_search, commands=["search"], state="*")

    @require_admin
    async def cb_search(call: types.CallbackQuery, state: FSMContext):
        """Начало поиска через callback"""
        await call.answer()
        await SearchRequest.waiting_query.set()
        await call.message.edit_text(
            "🔍 <b>Поиск заявок</b>\n\n"
            "Введите название компании, имя клиента или сферу деятельности:"
        )

    dp.register_callback_query_handler(cb_search, lambda c: c.data == CB_SEARCH, state="*")

    async def process_search(message: types.Message, state: FSMContext):
        """Обработка поискового запроса"""
        if message.text and message.text.startswith('/'):
            if message.text == '/cancel':
                await state.finish()
                return await message.answer("Поиск отменён.", reply_markup=admin_main_inline())

        query = message.text.strip()
        if len(query) < 2:
            return await message.answer("Введите минимум 2 символа для поиска.")

        results = search_requests(query, limit=20)

        await state.finish()

        if not results:
            await message.answer(
                f"🔍 По запросу «{e(query)}» ничего не найдено.\n\n"
                "Попробуйте другой запрос:",
                reply_markup=search_results_inline([])
            )
        else:
            await message.answer(
                f"🔍 Найдено: <b>{len(results)}</b>\n\n"
                f"Запрос: «{e(query)}»",
                reply_markup=search_results_inline(results)
            )

    dp.register_message_handler(process_search, state=SearchRequest.waiting_query)

    # ==================== МАССОВЫЕ ОПЕРАЦИИ ====================

    @require_admin
    async def cmd_mass_ops(message: types.Message):
        """Массовые операции"""
        await message.answer(
            "⚡ <b>Массовые операции</b>\n\n"
            "Выберите действие:",
            reply_markup=mass_ops_start_inline()
        )

    dp.register_message_handler(cmd_mass_ops, lambda m: m.text == BTN_MASS_OPS, state="*")
    dp.register_message_handler(cmd_mass_ops, commands=["mass_ops"], state="*")

    @require_admin
    async def cb_mass_ops(call: types.CallbackQuery):
        """Меню массовых операций"""
        await call.answer()
        await call.message.edit_text(
            "⚡ <b>Массовые операции</b>\n\n"
            "Выберите действие:",
            reply_markup=mass_ops_start_inline()
        )

    dp.register_callback_query_handler(cb_mass_ops, lambda c: c.data == CB_MASS_OPS, state="*")

    @require_admin
    async def cb_mass_archive(call: types.CallbackQuery, state: FSMContext):
        """Подготовка массового архивирования"""
        action_type = call.data.split("_")[-1]  # completed, old

        # Находим подходящие заявки
        all_requests = list_all_requests(0, 1000)

        to_archive = []
        if action_type == "completed":
            to_archive = [r for r in all_requests if r.get('status') == 'generated_ok']
            desc = "сгенерированных"
        elif action_type == "old":
            # Старше 30 дней
            from datetime import timedelta
            cutoff = datetime.now() - timedelta(days=30)
            to_archive = [
                r for r in all_requests
                if r.get('created_at') and r['created_at'] < cutoff
                and r.get('status') not in ('generating', 'queued')
            ]
            desc = "старых (30+ дней)"

        if not to_archive:
            return await call.answer(f"Нет заявок для архивации", show_alert=True)

        ids = [str(r['id']) for r in to_archive]
        await state.update_data(mass_ids=ids, mass_action="archive", mass_type=action_type)

        await call.message.edit_text(
            f"🗄 <b>Архивирование {desc}</b>\n\n"
            f"Будет архивировано: <b>{len(ids)}</b> заявок\n\n"
            "Подтвердить?",
            reply_markup=mass_ops_confirm_inline(f"archive_{action_type}", len(ids))
        )

    dp.register_callback_query_handler(
        cb_mass_archive,
        lambda c: c.data and c.data.startswith(CB_MASS_ARCHIVE),
        state="*"
    )

    @require_admin
    async def cb_mass_delete(call: types.CallbackQuery, state: FSMContext):
        """Подготовка массового удаления"""
        action_type = call.data.split("_")[-1]  # errors

        all_requests = list_all_requests(0, 1000)

        to_delete = []
        if action_type == "errors":
            to_delete = [r for r in all_requests if r.get('status') == 'generated_error']
            desc = "с ошибками генерации"

        if not to_delete:
            return await call.answer("Нет заявок для удаления", show_alert=True)

        ids = [str(r['id']) for r in to_delete]
        await state.update_data(mass_ids=ids, mass_action="delete", mass_type=action_type)

        await call.message.edit_text(
            f"🗑 <b>Удаление заявок {desc}</b>\n\n"
            f"Будет удалено: <b>{len(ids)}</b> заявок\n\n"
            "⚠️ Это действие нельзя отменить!\n"
            "Подтвердить?",
            reply_markup=mass_ops_confirm_inline(f"delete_{action_type}", len(ids))
        )

    dp.register_callback_query_handler(
        cb_mass_delete,
        lambda c: c.data and c.data.startswith(CB_MASS_DELETE),
        state="*"
    )

    @require_admin
    async def cb_mass_confirm(call: types.CallbackQuery, state: FSMContext):
        """Выполнение массовой операции"""
        data = await state.get_data()
        ids = data.get('mass_ids', [])
        action = data.get('mass_action')

        if not ids:
            await state.finish()
            return await call.answer("Нет заявок для обработки", show_alert=True)

        await call.answer("Выполняю...")

        admin = get_user_by_tgid(call.from_user.id)
        admin_id = str(admin['id']) if admin else None

        if action == "archive":
            count = mass_archive_requests(ids)
            log_activity(admin_id, "mass_archive", None, None, {"count": count})
            result_text = f"✅ Архивировано: {count} заявок"
        elif action == "delete":
            count = mass_delete_requests(ids)
            log_activity(admin_id, "mass_delete", None, None, {"count": count})
            result_text = f"✅ Удалено: {count} заявок"
        else:
            result_text = "❓ Неизвестное действие"

        await state.finish()
        await call.message.edit_text(result_text, reply_markup=admin_main_inline())

    dp.register_callback_query_handler(
        cb_mass_confirm,
        lambda c: c.data and c.data.startswith(CB_MASS_CONFIRM),
        state="*"
    )

    @require_admin
    async def cb_mass_cancel(call: types.CallbackQuery, state: FSMContext):
        """Отмена массовой операции"""
        await state.finish()
        await call.answer("Отменено")
        await call.message.edit_text("Операция отменена.", reply_markup=admin_main_inline())

    dp.register_callback_query_handler(cb_mass_cancel, lambda c: c.data == CB_MASS_CANCEL, state="*")

    # ==================== ЭКСПОРТ ====================

    @require_admin
    async def cmd_export(message: types.Message):
        """Меню экспорта"""
        await message.answer(
            "📊 <b>Экспорт статистики</b>\n\n"
            "Выберите формат:",
            reply_markup=export_options_inline()
        )

    dp.register_message_handler(cmd_export, lambda m: m.text == BTN_EXPORT, state="*")
    dp.register_message_handler(cmd_export, commands=["export"], state="*")

    @require_admin
    async def cb_export_excel(call: types.CallbackQuery):
        """Экспорт в Excel"""
        await call.answer("Формирую отчёт...")

        try:
            # Пробуем использовать openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.chart import BarChart, Reference

            wb = Workbook()

            # ===== Лист 1: Статистика по дням =====
            ws1 = wb.active
            ws1.title = "По дням"

            # Заголовки
            headers = ["Дата", "Всего", "Завершено", "В архиве", "В работе"]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Данные
            daily_stats = get_stats_for_export(30)
            for row_idx, stat in enumerate(daily_stats, 2):
                date_val = stat.get('date')
                if hasattr(date_val, 'strftime'):
                    date_val = date_val.strftime('%d.%m.%Y')
                ws1.cell(row=row_idx, column=1, value=date_val)
                ws1.cell(row=row_idx, column=2, value=stat.get('total_requests', 0))
                ws1.cell(row=row_idx, column=3, value=stat.get('completed', 0))
                ws1.cell(row=row_idx, column=4, value=stat.get('archived', 0))
                ws1.cell(row=row_idx, column=5, value=stat.get('pending', 0))

            # Диаграмма
            if len(daily_stats) > 0:
                chart = BarChart()
                chart.type = "col"
                chart.title = "Заявки по дням"
                chart.y_axis.title = "Количество"
                chart.x_axis.title = "Дата"

                data = Reference(ws1, min_col=2, min_row=1, max_col=5, max_row=len(daily_stats)+1)
                cats = Reference(ws1, min_col=1, min_row=2, max_row=len(daily_stats)+1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.shape = 4
                chart.width = 20
                chart.height = 10
                ws1.add_chart(chart, "G2")

            # ===== Лист 2: Статистика менеджеров =====
            ws2 = wb.create_sheet("Менеджеры")

            headers2 = ["Менеджер", "Всего", "Завершено", "За неделю", "Сегодня"]
            for col, header in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            managers_stats = get_managers_stats_for_export()
            for row_idx, stat in enumerate(managers_stats, 2):
                ws2.cell(row=row_idx, column=1, value=stat.get('manager_name', '—'))
                ws2.cell(row=row_idx, column=2, value=stat.get('total_requests', 0))
                ws2.cell(row=row_idx, column=3, value=stat.get('completed', 0))
                ws2.cell(row=row_idx, column=4, value=stat.get('this_week', 0))
                ws2.cell(row=row_idx, column=5, value=stat.get('today', 0))

            # Диаграмма менеджеров
            if len(managers_stats) > 0:
                chart2 = BarChart()
                chart2.type = "col"
                chart2.title = "Топ менеджеров"
                chart2.y_axis.title = "Заявок"

                data2 = Reference(ws2, min_col=2, min_row=1, max_col=3, max_row=min(len(managers_stats)+1, 11))
                cats2 = Reference(ws2, min_col=1, min_row=2, max_row=min(len(managers_stats)+1, 11))
                chart2.add_data(data2, titles_from_data=True)
                chart2.set_categories(cats2)
                chart2.width = 15
                chart2.height = 10
                ws2.add_chart(chart2, "G2")

            # ===== Лист 3: По статусам =====
            ws3 = wb.create_sheet("По статусам")

            headers3 = ["Статус", "Количество"]
            for col, header in enumerate(headers3, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            by_status = get_requests_by_status()
            for row_idx, stat in enumerate(by_status, 2):
                ws3.cell(row=row_idx, column=1, value=stat.get('status', '—'))
                ws3.cell(row=row_idx, column=2, value=stat.get('count', 0))

            # Авто-ширина колонок
            for ws in [ws1, ws2, ws3]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = max_length + 2

            # Сохраняем в буфер
            file_buffer = io.BytesIO()
            wb.save(file_buffer)
            file_buffer.seek(0)

            filename = f"statistics_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

            await call.message.answer_document(
                types.InputFile(file_buffer, filename=filename),
                caption=f"📊 Статистика за последние 30 дней\n\n"
                        f"Дата создания: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            )

        except ImportError:
            # Fallback на CSV если openpyxl не установлен
            await call.message.answer(
                "⚠️ Для Excel-экспорта требуется библиотека openpyxl.\n"
                "Установите: pip install openpyxl\n\n"
                "Пока могу экспортировать только в JSON."
            )

    dp.register_callback_query_handler(cb_export_excel, lambda c: c.data == CB_EXPORT_EXCEL, state="*")

    @require_admin
    async def cb_export_pdf(call: types.CallbackQuery):
        """Экспорт в PDF"""
        await call.answer("Формирую отчёт...")

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            # Буфер для PDF
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)

            elements = []
            styles = getSampleStyleSheet()

            # Заголовок
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center
            )
            elements.append(Paragraph("Statistics Report", title_style))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
            elements.append(Spacer(1, 20))

            # Общая статистика
            overall = get_overall_stats()
            elements.append(Paragraph("Overview", styles['Heading2']))

            overview_data = [
                ["Metric", "Value"],
                ["Total Users", str(overall.get('total_users', 0))],
                ["Total Managers", str(overall.get('total_managers', 0))],
                ["Total Requests", str(overall.get('total_requests', 0))],
                ["Today", str(overall.get('requests_today', 0))],
                ["This Week", str(overall.get('requests_this_week', 0))],
            ]

            t = Table(overview_data, colWidths=[8*cm, 4*cm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#D9E2F3')),
                ('GRID', (0, 0), (-1, -1), 1, colors.white),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 20))

            # По статусам
            elements.append(Paragraph("By Status", styles['Heading2']))

            by_status = get_requests_by_status()
            status_data = [["Status", "Count"]]
            for s in by_status:
                status_data.append([s.get('status', '—'), str(s.get('count', 0))])

            t2 = Table(status_data, colWidths=[8*cm, 4*cm])
            t2.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ]))
            elements.append(t2)
            elements.append(Spacer(1, 20))

            # Топ менеджеров
            elements.append(Paragraph("Top Managers", styles['Heading2']))

            managers_stats = get_managers_stats_for_export()[:10]
            mgr_data = [["Manager", "Total", "Completed", "This Week"]]
            for m in managers_stats:
                mgr_data.append([
                    m.get('manager_name', '—'),
                    str(m.get('total_requests', 0)),
                    str(m.get('completed', 0)),
                    str(m.get('this_week', 0))
                ])

            t3 = Table(mgr_data, colWidths=[6*cm, 3*cm, 3*cm, 3*cm])
            t3.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ]))
            elements.append(t3)

            # Генерация PDF
            doc.build(elements)
            buffer.seek(0)

            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

            await call.message.answer_document(
                types.InputFile(buffer, filename=filename),
                caption=f"📄 PDF-отчёт\n\n"
                        f"Дата создания: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            )

        except ImportError:
            await call.message.answer(
                "⚠️ Для PDF-экспорта требуется библиотека reportlab.\n"
                "Установите: pip install reportlab"
            )
        except Exception as e:
            log.exception("PDF export error")
            await call.message.answer(f"❌ Ошибка генерации PDF: {str(e)[:200]}")

    dp.register_callback_query_handler(cb_export_pdf, lambda c: c.data == CB_EXPORT_PDF, state="*")

