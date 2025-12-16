from typing import Optional, List
import io
import base64
import httpx
import logging
from datetime import datetime

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from config import settings
from routes.auth import get_admin_user
import db

log = logging.getLogger(__name__)

router = APIRouter()


class RejectRequest(BaseModel):
    reason: str


class BroadcastRequest(BaseModel):
    message: str
    photo: Optional[str] = None
    recipient_ids: Optional[List[str]] = None


class MassActionRequest(BaseModel):
    ids: List[str]
    confirmation_code: Optional[str] = None


class CreateGroupRequest(BaseModel):
    name: str
    description: Optional[str] = None


class AddToGroupRequest(BaseModel):
    user_id: str
    role: Optional[str] = "member"


class CreateInviteCodeRequest(BaseModel):
    group_id: Optional[str] = None
    name: Optional[str] = None
    max_uses: Optional[int] = None
    expires_in_days: Optional[int] = None
    auto_approve: Optional[bool] = False
    notes: Optional[str] = None


class UpdateInviteCodeRequest(BaseModel):
    name: Optional[str] = None
    max_uses: Optional[int] = None
    auto_approve: Optional[bool] = None
    is_active: Optional[bool] = None
    notes: Optional[str] = None
    group_id: Optional[str] = None


# ==================== Dashboard ====================

@router.get("/dashboard")
async def get_dashboard(user: dict = Depends(get_admin_user)):
    """Get admin dashboard data."""
    return await db.get_dashboard_stats()


# ==================== Managers ====================

@router.get("/managers")
async def list_managers(user: dict = Depends(get_admin_user)):
    """List managers. If admin is in a group, show only group members."""
    # Try to get group-based managers first
    managers = await db.list_managers_by_admin(str(user['id']))
    if not managers:
        # Fallback to all managers
        managers = await db.list_managers()
    return managers


@router.get("/managers/{manager_id}")
async def get_manager(manager_id: str, user: dict = Depends(get_admin_user)):
    """Get manager details. Only accessible if manager belongs to admin's groups."""
    manager = await db.get_user_by_id(manager_id)
    if not manager:
        raise HTTPException(status_code=404, detail="Manager not found")

    # Check access
    if not await db.is_manager_accessible_by_admin(manager_id, str(user['id'])):
        raise HTTPException(status_code=403, detail="Доступ запрещен. Вы можете просматривать только своих менеджеров")

    stats = await db.get_user_stats(manager_id)
    return {**manager, "id": str(manager["id"]), "stats": stats}


@router.post("/managers/{manager_id}/block")
async def block_manager(manager_id: str, user: dict = Depends(get_admin_user)):
    """Block a manager. Only accessible if manager belongs to admin's groups."""
    # Check access
    if not await db.is_manager_accessible_by_admin(manager_id, str(user['id'])):
        raise HTTPException(status_code=403, detail="Доступ запрещен. Вы можете блокировать только своих менеджеров")

    await db.block_user(manager_id)
    return {"success": True}


@router.post("/managers/{manager_id}/unblock")
async def unblock_manager(manager_id: str, user: dict = Depends(get_admin_user)):
    """Unblock a manager. Only accessible if manager belongs to admin's groups."""
    # Check access
    if not await db.is_manager_accessible_by_admin(manager_id, str(user['id'])):
        raise HTTPException(status_code=403, detail="Доступ запрещен. Вы можете разблокировать только своих менеджеров")

    await db.unblock_user(manager_id)
    return {"success": True}


@router.delete("/managers/{manager_id}")
async def delete_manager(
    manager_id: str,
    user: dict = Depends(get_admin_user),
    confirmation: Optional[str] = None
):
    """Delete a manager with anti-nuke protection. Only accessible if manager belongs to admin's groups."""
    admin_id = str(user['id'])

    # Check access
    if not await db.is_manager_accessible_by_admin(manager_id, admin_id):
        raise HTTPException(status_code=403, detail="Доступ запрещен. Вы можете удалять только своих менеджеров")

    # Anti-nuke check
    can_delete, error_msg = await db.can_delete_manager(admin_id, manager_id)
    if not can_delete:
        raise HTTPException(status_code=429, detail=error_msg)

    # Log the deletion
    await db.log_deletion(
        action_type='delete_manager',
        target_type='user',
        target_id=manager_id,
        performed_by=admin_id,
        reason=confirmation or 'Admin deletion'
    )

    await db.delete_user(manager_id)
    return {"success": True}


@router.post("/managers/{manager_id}/make-admin")
async def make_manager_admin(manager_id: str, user: dict = Depends(get_admin_user)):
    """Promote manager to admin role."""
    await db.update_user_role(manager_id, 'admin')
    return {"success": True}


# ==================== Pending Registrations ====================

@router.get("/pending")
async def list_pending(user: dict = Depends(get_admin_user)):
    """List pending registrations from admin's groups."""
    return await db.list_pending_registrations(admin_id=str(user['id']))


@router.post("/pending/{user_id}/approve")
async def approve_registration(user_id: str, user: dict = Depends(get_admin_user)):
    """Approve a pending registration."""
    # Check current status first
    current_status = await db.get_user_approval_status(user_id)
    if not current_status:
        raise HTTPException(status_code=404, detail="User not found")

    if current_status == 'approved':
        return {"success": True, "message": "Already approved"}

    if current_status == 'rejected':
        raise HTTPException(status_code=400, detail="User was already rejected by another admin")

    result = await db.approve_user(user_id)
    if not result:
        raise HTTPException(status_code=400, detail="Status already changed by another admin")

    return {"success": True}


@router.post("/pending/{user_id}/reject")
async def reject_registration(
    user_id: str,
    data: RejectRequest,
    user: dict = Depends(get_admin_user)
):
    """Reject a pending registration. Only accessible if user belongs to admin's groups."""
    # Check current status first
    current_status = await db.get_user_approval_status(user_id)
    if not current_status:
        raise HTTPException(status_code=404, detail="User not found")

    # Check if user is a manager and if they belong to admin's groups
    user_data = await db.get_user_by_id(user_id)
    if user_data and user_data.get('role') == 'manager':
        # Check if user is registered via invite code that belongs to admin's groups
        if user_data.get('registered_via_code'):
            async with await db.get_conn() as conn:
                async with conn.cursor(row_factory=dict_row) as cur:
                    await cur.execute(
                        """SELECT ic.group_id, ag.created_by
                           FROM invite_codes ic
                           LEFT JOIN admin_groups ag ON ag.id = ic.group_id
                           WHERE ic.id = %s""",
                        (user_data['registered_via_code'],)
                    )
                    invite_data = await cur.fetchone()
                    if invite_data and invite_data.get('created_by'):
                        if str(invite_data['created_by']) != str(user['id']):
                            raise HTTPException(status_code=403, detail="Доступ запрещен. Вы можете отклонять только менеджеров из своих групп")

    if current_status == 'rejected':
        return {"success": True, "message": "Already rejected"}

    if current_status == 'approved':
        raise HTTPException(status_code=400, detail="User was already approved by another admin")

    result = await db.reject_user(user_id, data.reason)
    if not result:
        raise HTTPException(status_code=400, detail="Status already changed by another admin")

    return {"success": True}


@router.post("/pending/{user_id}/reset")
async def reset_registration(user_id: str, user: dict = Depends(get_admin_user)):
    """Reset a rejected user's status to pending (allow re-application)."""
    current_status = await db.get_user_approval_status(user_id)
    if not current_status:
        raise HTTPException(status_code=404, detail="User not found")

    if current_status != 'rejected':
        raise HTTPException(status_code=400, detail="Can only reset rejected users")

    await db.reset_user_approval(user_id)
    return {"success": True}


# ==================== Requests ====================

@router.get("/requests")
async def list_all_requests(
    status: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    user: dict = Depends(get_admin_user)
):
    """List all requests (admin view)."""
    offset = (page - 1) * limit
    requests = await db.list_requests(status=status, limit=limit, offset=offset)

    return {
        "items": [{**r, "id": str(r["id"])} for r in requests],
        "page": page,
        "limit": limit,
    }


@router.get("/requests/search")
async def search_requests(q: str, user: dict = Depends(get_admin_user)):
    """Search requests by company or client name."""
    # Simple search - can be enhanced with full-text search
    all_requests = await db.list_requests(limit=1000)
    q_lower = q.lower()

    results = [
        {**r, "id": str(r["id"])}
        for r in all_requests
        if q_lower in (r.get('company_name', '') or '').lower() or
           q_lower in (r.get('client_name', '') or '').lower()
    ]

    return {"items": results[:50]}


@router.post("/requests/mass-archive")
async def mass_archive(data: MassActionRequest, user: dict = Depends(get_admin_user)):
    """Archive multiple requests."""
    await db.mass_archive_requests(data.ids)
    return {"success": True, "count": len(data.ids)}


@router.post("/requests/mass-delete")
async def mass_delete(data: MassActionRequest, user: dict = Depends(get_admin_user)):
    """Delete multiple requests with anti-nuke protection."""
    admin_id = str(user['id'])
    count = len(data.ids)

    # Anti-nuke check
    can_delete, error_msg = await db.can_mass_delete_requests(admin_id, count)
    if not can_delete:
        raise HTTPException(status_code=429, detail=error_msg)

    # Require confirmation code for large deletions
    require_confirm = int(await db.get_anti_nuke_setting('require_confirmation_above') or '5')
    if count > require_confirm:
        expected_code = f"DELETE-{count}"
        if data.confirmation_code != expected_code:
            raise HTTPException(
                status_code=400,
                detail=f"Для удаления {count} заявок требуется код подтверждения: {expected_code}"
            )

    # Log the deletion
    await db.log_deletion(
        action_type='mass_delete_requests',
        target_type='request',
        target_ids=data.ids,
        performed_by=admin_id,
        reason=f'Mass deletion of {count} requests'
    )

    await db.mass_delete_requests(data.ids)
    return {"success": True, "count": count}


# ==================== Admin Groups ====================

@router.get("/groups")
async def list_groups(user: dict = Depends(get_admin_user)):
    """List admin groups created by the current admin."""
    return await db.list_admin_groups(created_by=str(user['id']))


@router.post("/groups")
async def create_group(data: CreateGroupRequest, user: dict = Depends(get_admin_user)):
    """Create a new admin group."""
    group = await db.create_admin_group(
        name=data.name,
        description=data.description,
        created_by=str(user['id'])
    )
    return {**group, "id": str(group["id"])}


@router.get("/groups/{group_id}")
async def get_group(group_id: str, user: dict = Depends(get_admin_user)):
    """Get admin group details with members."""
    group = await db.get_admin_group(group_id)
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    return group


@router.post("/groups/{group_id}/members")
async def add_group_member(
    group_id: str,
    data: AddToGroupRequest,
    user: dict = Depends(get_admin_user)
):
    """Add a user to an admin group. Only accessible if group was created by admin."""
    group = await db.get_admin_group(group_id)
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    if str(group['created_by']) != str(user['id']):
        raise HTTPException(status_code=403, detail="Доступ запрещен. Вы можете добавлять пользователей только в свои группы")

    membership = await db.add_user_to_group(
        user_id=data.user_id,
        group_id=group_id,
        role=data.role,
        added_by=str(user['id'])
    )
    return {"success": True, "membership": membership}


@router.delete("/groups/{group_id}/members/{user_id}")
async def remove_group_member(
    group_id: str,
    user_id: str,
    user: dict = Depends(get_admin_user)
):
    """Remove a user from an admin group. Only accessible if group was created by admin."""
    group = await db.get_admin_group(group_id)
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    if str(group['created_by']) != str(user['id']):
        raise HTTPException(status_code=403, detail="Доступ запрещен. Вы можете удалять пользователей только из своих групп")

    await db.remove_user_from_group(user_id, group_id)
    return {"success": True}


@router.get("/my-groups")
async def get_my_groups(user: dict = Depends(get_admin_user)):
    """Get groups the current admin belongs to."""
    return await db.get_user_groups(str(user['id']))


# ==================== Invite Codes ====================

@router.get("/invite-codes")
async def list_invite_codes(
    group_id: Optional[str] = None,
    user: dict = Depends(get_admin_user)
):
    """List invite codes created by this admin or for their groups."""
    codes = await db.list_invite_codes(created_by=str(user['id']), group_id=group_id)
    return [
        {**c, "id": str(c["id"]), "group_id": str(c["group_id"]) if c.get("group_id") else None}
        for c in codes
    ]


@router.post("/invite-codes")
async def create_invite_code(data: CreateInviteCodeRequest, user: dict = Depends(get_admin_user)):
    """Create a new invite code. Group is required."""
    from datetime import timedelta

    # Group is required
    if not data.group_id:
        raise HTTPException(status_code=400, detail="Группа обязательна для создания инвайт-кода")

    # Verify that the group belongs to this admin
    group = await db.get_admin_group(data.group_id)
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена")

    if str(group['created_by']) != str(user['id']):
        raise HTTPException(status_code=403, detail="Вы можете создавать инвайт-коды только для своих групп")

    expires_at = None
    if data.expires_in_days:
        expires_at = datetime.now() + timedelta(days=data.expires_in_days)

    code = await db.create_invite_code(
        created_by=str(user['id']),
        group_id=data.group_id,
        name=data.name,
        max_uses=data.max_uses,
        expires_at=expires_at,
        auto_approve=data.auto_approve or False,
        notes=data.notes
    )

    return {
        **code,
        "id": str(code["id"]),
        "group_id": str(code["group_id"]) if code.get("group_id") else None
    }


@router.get("/invite-codes/{code_id}")
async def get_invite_code_details(code_id: str, user: dict = Depends(get_admin_user)):
    """Get invite code details with usage info."""
    code = await db.get_invite_code_by_id(code_id)
    if not code:
        raise HTTPException(status_code=404, detail="Invite code not found")

    # Check if user owns this code or is superadmin
    if str(code['created_by']) != str(user['id']):
        raise HTTPException(status_code=403, detail="Access denied")

    usage = await db.get_invite_code_usage(code_id)

    return {
        **code,
        "id": str(code["id"]),
        "group_id": str(code["group_id"]) if code.get("group_id") else None,
        "usage": [
            {**u, "id": str(u["id"])}
            for u in usage
        ]
    }


@router.patch("/invite-codes/{code_id}")
async def update_invite_code(
    code_id: str,
    data: UpdateInviteCodeRequest,
    user: dict = Depends(get_admin_user)
):
    """Update an invite code. Only accessible if code belongs to admin."""
    code = await db.get_invite_code_by_id(code_id)
    if not code:
        raise HTTPException(status_code=404, detail="Invite code not found")

    if str(code['created_by']) != str(user['id']):
        raise HTTPException(status_code=403, detail="Доступ запрещен. Вы можете редактировать только свои инвайт-коды")

    # If updating group_id, verify it belongs to admin
    if data.group_id and data.group_id != code.get('group_id'):
        group = await db.get_admin_group(data.group_id)
        if not group:
            raise HTTPException(status_code=404, detail="Группа не найдена")
        if str(group['created_by']) != str(user['id']):
            raise HTTPException(status_code=403, detail="Доступ запрещен. Вы можете использовать только свои группы")

    update_data = data.model_dump(exclude_none=True)
    updated = await db.update_invite_code(code_id, update_data)

    return {
        **updated,
        "id": str(updated["id"]),
        "group_id": str(updated["group_id"]) if updated.get("group_id") else None
    }


@router.delete("/invite-codes/{code_id}")
async def delete_invite_code(code_id: str, user: dict = Depends(get_admin_user)):
    """Delete an invite code. Only accessible if code belongs to admin."""
    code = await db.get_invite_code_by_id(code_id)
    if not code:
        raise HTTPException(status_code=404, detail="Invite code not found")

    if str(code['created_by']) != str(user['id']):
        raise HTTPException(status_code=403, detail="Доступ запрещен. Вы можете удалять только свои инвайт-коды")

    await db.delete_invite_code(code_id)
    return {"success": True}


@router.get("/invite-codes/validate/{code}")
async def validate_invite_code(code: str):
    """Validate an invite code (public endpoint for registration)."""
    is_valid, result = await db.validate_invite_code(code)

    if not is_valid:
        return {"valid": False, "error": result}

    invite = result
    return {
        "valid": True,
        "group_name": invite.get('group_name'),
        "auto_approve": invite.get('auto_approve', False),
        "creator_name": f"{invite.get('creator_first_name', '')} {invite.get('creator_last_name', '')}".strip()
    }


# ==================== Broadcast ====================

async def send_telegram_message(tg_id: int, text: str, photo_data: Optional[str] = None) -> bool:
    """Send message to user via Telegram Bot API."""
    try:
        async with httpx.AsyncClient() as client:
            if photo_data:
                # Check if it's a base64 data URL
                if photo_data.startswith('data:'):
                    # Extract base64 content
                    # Format: data:image/jpeg;base64,/9j/4AAQ...
                    try:
                        header, b64_content = photo_data.split(',', 1)
                        image_bytes = base64.b64decode(b64_content)

                        # Determine file extension from header
                        ext = 'jpg'
                        if 'png' in header:
                            ext = 'png'
                        elif 'gif' in header:
                            ext = 'gif'
                        elif 'webp' in header:
                            ext = 'webp'

                        # Send as multipart/form-data
                        files = {
                            'photo': (f'photo.{ext}', image_bytes, f'image/{ext}')
                        }
                        data = {
                            'chat_id': str(tg_id),
                            'caption': text,
                            'parse_mode': 'HTML'
                        }

                        response = await client.post(
                            f"https://api.telegram.org/bot{settings.BOT_TOKEN}/sendPhoto",
                            data=data,
                            files=files,
                            timeout=30.0
                        )
                    except Exception as e:
                        log.error(f"Failed to decode base64 image: {e}")
                        # Fallback to text-only message
                        response = await client.post(
                            f"https://api.telegram.org/bot{settings.BOT_TOKEN}/sendMessage",
                            json={
                                "chat_id": tg_id,
                                "text": text,
                                "parse_mode": "HTML"
                            },
                            timeout=10.0
                        )
                else:
                    # It's a regular URL
                    response = await client.post(
                        f"https://api.telegram.org/bot{settings.BOT_TOKEN}/sendPhoto",
                        json={
                            "chat_id": tg_id,
                            "photo": photo_data,
                            "caption": text,
                            "parse_mode": "HTML"
                        },
                        timeout=10.0
                    )
            else:
                # Send text message
                response = await client.post(
                    f"https://api.telegram.org/bot{settings.BOT_TOKEN}/sendMessage",
                    json={
                        "chat_id": tg_id,
                        "text": text,
                        "parse_mode": "HTML"
                    },
                    timeout=10.0
                )

            result = response.json()
            if not result.get('ok'):
                log.warning(f"Failed to send message to {tg_id}: {result}")
                return False
            return True
    except Exception as e:
        log.error(f"Error sending message to {tg_id}: {e}")
        return False


@router.post("/broadcast")
async def send_broadcast(data: BroadcastRequest, user: dict = Depends(get_admin_user)):
    """Send broadcast message to managers."""
    if not data.message:
        raise HTTPException(status_code=400, detail="Message is required")

    # Get recipients
    if data.recipient_ids:
        # Get specific managers by ID
        managers = await db.list_managers()
        recipients = [m for m in managers if str(m['id']) in data.recipient_ids and not m.get('is_blocked')]
    else:
        # Send to all active managers
        managers = await db.list_managers()
        recipients = [m for m in managers if not m.get('is_blocked')]

    if not recipients:
        return {
            "success": False,
            "error": "No recipients found",
            "sent_count": 0,
            "failed_count": 0,
        }

    # Send messages
    sent_count = 0
    failed_count = 0

    for manager in recipients:
        tg_id = manager.get('tg_id')
        if not tg_id:
            failed_count += 1
            continue

        success = await send_telegram_message(tg_id, data.message, data.photo)
        if success:
            sent_count += 1
        else:
            failed_count += 1

    log.info(f"Broadcast sent: {sent_count} success, {failed_count} failed")

    return {
        "success": True,
        "sent_count": sent_count,
        "failed_count": failed_count,
        "total_recipients": len(recipients),
    }


# ==================== Stats ====================

@router.get("/stats/overview")
async def get_stats_overview(user: dict = Depends(get_admin_user)):
    """Get overview statistics."""
    return await db.get_overview_stats()


@router.get("/stats/by-status")
async def get_stats_by_status(user: dict = Depends(get_admin_user)):
    """Get statistics by status."""
    return await db.get_stats_by_status()


@router.get("/stats/by-day")
async def get_stats_by_day(days: int = 7, user: dict = Depends(get_admin_user)):
    """Get statistics by day."""
    return await db.get_stats_by_day(days)


@router.get("/stats/managers")
async def get_manager_stats(user: dict = Depends(get_admin_user)):
    """Get manager statistics."""
    return await db.get_manager_stats()


# ==================== Export ====================

@router.get("/export/excel")
async def export_excel(user: dict = Depends(get_admin_user)):
    """Export statistics to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statistics"

    # Get data
    stats = await db.get_overview_stats()
    by_status = await db.get_stats_by_status()
    managers = await db.get_manager_stats()

    # Overview
    ws['A1'] = 'AutoSites Statistics'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'

    ws['A4'] = 'Overview'
    ws['A4'].font = Font(bold=True)
    ws['A5'] = 'Total Requests'
    ws['B5'] = stats.get('total_requests', 0)
    ws['A6'] = 'Total Managers'
    ws['B6'] = stats.get('total_managers', 0)
    ws['A7'] = 'This Month'
    ws['B7'] = stats.get('this_month', 0)

    # By Status
    ws['A9'] = 'By Status'
    ws['A9'].font = Font(bold=True)
    row = 10
    for item in by_status:
        ws[f'A{row}'] = item.get('status', 'unknown')
        ws[f'B{row}'] = item.get('count', 0)
        row += 1

    # Top Managers
    ws['A{}'.format(row + 1)] = 'Top Managers'
    ws['A{}'.format(row + 1)].font = Font(bold=True)
    row += 2
    ws[f'A{row}'] = 'Name'
    ws[f'B{row}'] = 'Username'
    ws[f'C{row}'] = 'Requests'
    row += 1
    for m in managers[:10]:
        ws[f'A{row}'] = f"{m.get('first_name', '')} {m.get('last_name', '')}"
        ws[f'B{row}'] = f"@{m.get('username', '')}" if m.get('username') else '-'
        ws[f'C{row}'] = m.get('request_count', 0)
        row += 1

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=stats_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )


@router.get("/export/pdf")
async def export_pdf(user: dict = Depends(get_admin_user)):
    """Export statistics to PDF."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab not installed")

    # Get data
    stats = await db.get_overview_stats()
    by_status = await db.get_stats_by_status()
    managers = await db.get_manager_stats()

    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("AutoSites Statistics", styles['Title']))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Overview
    elements.append(Paragraph("Overview", styles['Heading2']))
    overview_data = [
        ['Metric', 'Value'],
        ['Total Requests', str(stats.get('total_requests', 0))],
        ['Total Managers', str(stats.get('total_managers', 0))],
        ['This Month', str(stats.get('this_month', 0))],
    ]
    overview_table = Table(overview_data, colWidths=[200, 100])
    overview_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(overview_table)
    elements.append(Spacer(1, 20))

    # By Status
    elements.append(Paragraph("By Status", styles['Heading2']))
    status_data = [['Status', 'Count']]
    for item in by_status:
        status_data.append([item.get('status', 'unknown'), str(item.get('count', 0))])
    status_table = Table(status_data, colWidths=[200, 100])
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(status_table)
    elements.append(Spacer(1, 20))

    # Top Managers
    elements.append(Paragraph("Top Managers", styles['Heading2']))
    manager_data = [['Name', 'Username', 'Requests']]
    for m in managers[:10]:
        manager_data.append([
            f"{m.get('first_name', '')} {m.get('last_name', '')}",
            f"@{m.get('username', '')}" if m.get('username') else '-',
            str(m.get('request_count', 0))
        ])
    manager_table = Table(manager_data, colWidths=[150, 100, 80])
    manager_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(manager_table)

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=stats_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )

